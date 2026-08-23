"""UK Office for National Statistics — the GDP real-time database.

ONS publishes a genuine real-time database as a workbook: one column
per publication vintage, going back to 1961. For real GDP (``ABMI``,
chained volume measures) that is roughly 750 vintages — by far the
deepest single-country archive wired up here.

    https://www.ons.gov.uk/economy/grossdomesticproductgdp/datasets/
        realtimedatabaseforukgdpabmi

HOW THE VINTAGE IS IDENTIFIED — AND WHAT IT IS NOT
--------------------------------------------------
Row 4 of each data sheet holds hand-typed labels of the form::

    Jun-18 [2016 prices]\\nQNA
    Aug-18\\n1st

so a vintage is identified by **publication month plus release stage**
(``1st`` first estimate, ``M1``/``M2`` preliminary and second estimate,
``QNA`` Quarterly National Accounts) — *not* by an exact release day.
Several stages land in the same month, so month alone does not
identify a column.

The tidy schema needs one timestamp per edition, so this connector
sets the vintage to the first of the publication month and, where a
month carries several stages, adds one day per additional stage **in
the workbook's own left-to-right (chronological) order**. That day
component is an ordering device, not a release date. Use it to order
editions; do not use it to date a release event. The stage tag itself
is preserved by :func:`parse_ons_realtime_workbook`, which returns it
as a fourth column.

The bracketed ``[YYYY prices]`` annotations mark chained-volume
rebasings. Levels either side of one are not comparable — which the
growth-rate default in the revision tests handles, since a rebasing
rescales a whole column and cancels out of that column's growth rates.

HEADER GRAMMAR IS MESSY, ON PURPOSE OF NOBODY'S
------------------------------------------------
The labels are typed by hand and contain: a ``\\n`` separator that is
sometimes a plain space, four-letter months (``June-26``), and a
handful of outright typos (``Feb-772``). Columns whose label cannot be
parsed are skipped and counted rather than guessed at; the count is
returned in the frame's ``attrs``.
"""
from __future__ import annotations

import io
import json
import re

import pandas as pd

from ..._http import safe_get_bytes, safe_get_bytes_cached
from ._base import VintagePanel, normalize_vintage_frame, register_provider
from .catalog import SeriesSpec, register_catalog


ONS_BASE = "https://www.ons.gov.uk"
ONS_DATASET_INDEX = ONS_BASE + "/economy/grossdomesticproductgdp/datasets/{slug}/data"
ONS_FILE_URL = ONS_BASE + "/file?uri={uri}/{filename}"

_UA = "puremacro (real-time vintage reader)"

#: Canonical variable -> (dataset slug, units).
ONS_REALTIME_DATASETS: dict[str, tuple[str, str]] = {
    "gdp_real": ("realtimedatabaseforukgdpabmi", "level"),
    "gdp_nom": ("realtimedatabaseforukgdpybha", "level"),
}

# "Q1 1955" — quarter first, unlike almost every other source here.
_PERIOD_RE = re.compile(r"^\s*Q([1-4])\s+(\d{4})\s*$", re.I)

# "Jun-18 [2016 prices]\nQNA" / "Aug-18 1st" / "June-26\nQNA"
_VINTAGE_RE = re.compile(
    r"^\s*([A-Za-z]{3,9})\s*-\s*(\d{2,4})"      # month - year
    r"(?:\s*\[[^\]]*\])?"                        # optional [YYYY prices]
    r"[\s\n]*(.*?)\s*$",                         # optional stage tag
    re.S,
)

_MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}
_MONTHS.update({m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)})
_MONTHS["sept"] = 9
_MONTHS["june"] = 6
_MONTHS["july"] = 7

#: Two-digit years below this belong to the 2000s. The ONS archive
#: starts in 1961 and the sheets are named by vintage era, but the
#: pivot must NOT be derived from the sheet name: a '[1975 prices]'
#: annotation inside the '1983 - 2003' sheet would then flip a
#: 'Dec-82' column into 2082.
_YEAR_PIVOT = 60


def _parse_period(label) -> pd.Timestamp | None:
    m = _PERIOD_RE.match(str(label or ""))
    if not m:
        return None
    quarter, year = m.groups()
    return pd.Period(f"{year}Q{quarter}", freq="Q").to_timestamp()


def parse_ons_vintage_label(label) -> tuple[pd.Timestamp | None, str]:
    """Parse an ONS column header into ``(month start, stage tag)``.

    Returns ``(None, "")`` for a header that cannot be read — the
    handful of genuine typos in the workbook — so the caller can count
    and report them instead of inventing a date.
    """
    m = _VINTAGE_RE.match(str(label or "").replace(" ", " "))
    if not m:
        return None, ""
    mon, year, stage = m.groups()
    month = _MONTHS.get(mon.strip().lower())
    if month is None:
        return None, ""
    y = int(year)
    if len(year) <= 2:
        y = 2000 + y if y < _YEAR_PIVOT else 1900 + y
    elif not (1900 <= y <= 2100):
        return None, ""
    try:
        return pd.Timestamp(y, month, 1), " ".join(stage.split())
    except ValueError:
        return None, ""


def parse_ons_realtime_workbook(raw: bytes) -> pd.DataFrame:
    """Parse an ONS real-time workbook into ``[date, vintage, value, stage]``.

    Pure function apart from the lazy ``openpyxl`` import. Every data
    sheet is read (they partition the vintages by era, not the periods)
    and concatenated.

    Where one publication month carries several release stages, the
    vintage timestamps are spread across successive days of that month
    in workbook column order, purely to keep editions distinct and
    correctly ordered — see the module docstring.
    """
    try:
        import openpyxl
    except ImportError as exc:                            # pragma: no cover
        raise ImportError(
            "reading the ONS real-time workbook needs openpyxl. Install it "
            "with `pip install openpyxl` (it ships in puremacro's 'dev' "
            "extra). Every other real-time provider here is openpyxl-free."
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True,
                                data_only=True)
    records: list[tuple] = []
    unparsed_headers: list[str] = []

    for sheet in wb.sheetnames:
        if sheet in ("Cover_sheet", "Table_of_contents"):
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = None
        for _ in range(4):                    # header is row 4
            try:
                header = next(rows)
            except StopIteration:
                header = None
                break
        if not header:
            continue

        # Column index -> (vintage month, stage), skipping unreadable ones.
        cols: dict[int, tuple[pd.Timestamp, str]] = {}
        for i, cell in enumerate(header[1:], start=1):
            if cell in (None, ""):
                continue
            ts, stage = parse_ons_vintage_label(cell)
            if ts is None:
                unparsed_headers.append(str(cell))
                continue
            cols[i] = (ts, stage)

        # Spread same-month vintages across successive days, in the
        # workbook's own left-to-right (chronological) order.
        seen: dict[pd.Timestamp, int] = {}
        resolved: dict[int, pd.Timestamp] = {}
        for i in sorted(cols):
            month_start, _stage = cols[i]
            k = seen.get(month_start, 0)
            seen[month_start] = k + 1
            resolved[i] = month_start + pd.Timedelta(days=k)

        for row in rows:
            if not row:
                continue
            date = _parse_period(row[0])
            if date is None:
                continue
            for i, (_month, stage) in cols.items():
                if i >= len(row):
                    continue
                val = row[i]
                if val in (None, ""):
                    continue
                try:
                    value = float(val)
                except (TypeError, ValueError):
                    continue
                records.append((date, resolved[i], value, stage))

    wb.close()
    if not records:
        out = pd.DataFrame(columns=["date", "vintage", "value", "stage"])
    else:
        out = (pd.DataFrame(records,
                            columns=["date", "vintage", "value", "stage"])
               .drop_duplicates(subset=["date", "vintage"], keep="last")
               .sort_values(["date", "vintage"]).reset_index(drop=True))
    out.attrs["unparsed_headers"] = unparsed_headers
    return out


def _current_workbook_url(slug: str, *, timeout: float, use_cache: bool) -> str:
    """Resolve the newest edition's download URL.

    The filename is not stable across editions (it even changes
    extension for pre-2018 ones), so it must be read from the edition's
    JSON rather than assumed.
    """
    getter = safe_get_bytes_cached if use_cache else safe_get_bytes
    index = json.loads(getter(ONS_DATASET_INDEX.format(slug=slug), timeout,
                              user_agent=_UA))
    datasets = index.get("datasets") or []
    if not datasets:
        raise RuntimeError(f"ONS dataset index for {slug!r} listed no editions")
    uri = datasets[0]["uri"]                    # newest first
    edition = json.loads(getter(ONS_BASE + uri + "/data", timeout,
                                user_agent=_UA))
    downloads = edition.get("downloads") or []
    if not downloads:
        raise RuntimeError(f"ONS edition {uri!r} listed no downloads")
    return ONS_FILE_URL.format(uri=uri, filename=downloads[0]["file"])


def fetch_ons_vintages(
    variable: str = "gdp_real",
    *,
    timeout: float = 180.0,
    use_cache: bool = True,
) -> pd.DataFrame:
    """Long ``[date, vintage, value, stage]`` for one UK variable."""
    if variable not in ONS_REALTIME_DATASETS:
        raise ValueError(
            f"no ONS real-time database mapped for {variable!r}; available: "
            f"{sorted(ONS_REALTIME_DATASETS)}"
        )
    slug, _units = ONS_REALTIME_DATASETS[variable]
    url = _current_workbook_url(slug, timeout=timeout, use_cache=use_cache)
    getter = safe_get_bytes_cached if use_cache else safe_get_bytes
    return parse_ons_realtime_workbook(getter(url, timeout, user_agent=_UA))


def fetch_ons_panel(
    countries, variables, *, timeout: float = 180.0, use_cache: bool = True,
    **_ignored,
) -> VintagePanel:
    """Registry entry point."""
    frames, failed = [], {}
    for country in countries:
        if str(country).upper() != "GBR":
            continue
        for variable in variables:
            if variable not in ONS_REALTIME_DATASETS:
                continue
            slug, units = ONS_REALTIME_DATASETS[variable]
            try:
                long = fetch_ons_vintages(
                    variable, timeout=timeout, use_cache=use_cache)
            except Exception as exc:
                failed[f"{country}:{variable}"] = f"{type(exc).__name__}: {exc}"
                continue
            frames.append(normalize_vintage_frame(
                long, country="GBR", variable=variable, provider="ons",
                series_id=slug, units=units,
            ))
    df = (pd.concat(frames, ignore_index=True) if frames
          else pd.DataFrame(columns=[
              "country", "variable", "date", "vintage", "value", "provider",
              "series_id", "units"]))
    return VintagePanel(df=df, metadata={"provider": "ons",
                                         "failed": failed})


def _register() -> None:
    register_catalog("ons", {
        "GBR": {var: SeriesSpec(slug, units, "ons",
                                "ONS real-time database workbook; vintage is "
                                "publication month + release stage")
                for var, (slug, units) in ONS_REALTIME_DATASETS.items()}
    })
    register_provider("ons", fetch_ons_panel, ["GBR"])


__all__ = [
    "ONS_BASE",
    "ONS_REALTIME_DATASETS",
    "parse_ons_realtime_workbook",
    "parse_ons_vintage_label",
    "fetch_ons_vintages",
    "fetch_ons_panel",
]
